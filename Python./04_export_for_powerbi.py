"""
04_export_for_powerbi.py
-------------------------
Exports the cleaned star-schema tables into:
  1. data/powerbi_data_model.xlsx  — one formatted sheet per table, ready for
     Power BI's "Get Data > Excel Workbook", with relationships documented
     on a cover sheet.
  2. data/csv_for_powerbi/*.csv    — the same tables as plain CSVs, for
     "Get Data > Text/CSV" if preferred.
"""
import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DB = "data/retail_ecommerce.db"
XLSX_OUT = "data/powerbi_data_model.xlsx"
CSV_DIR = "data/csv_for_powerbi"

NAVY = "1B2A4A"
GOLD = "E8A33D"

conn = sqlite3.connect(DB)
tables = {
    "dim_customers": pd.read_sql("SELECT * FROM dim_customers", conn),
    "dim_products": pd.read_sql("SELECT * FROM dim_products", conn),
    "dim_date": pd.read_sql("SELECT * FROM dim_date", conn),
    "dim_channel": pd.read_sql("SELECT * FROM dim_channel", conn),
    "fact_sales": pd.read_sql("SELECT * FROM fact_sales", conn),
}
delivered = pd.read_sql(
    "SELECT customer_id, order_id, order_date, sales_amount FROM fact_sales WHERE order_status='Delivered'",
    conn, parse_dates=["order_date"],
)
conn.close()

# ---- enrich dim_customers with a precomputed RFM snapshot for Power BI ----
AS_OF = pd.Timestamp("2026-06-30")
rfm = delivered.groupby("customer_id").agg(
    recency_days=("order_date", lambda s: (AS_OF - s.max()).days),
    frequency=("order_id", "nunique"),
    monetary=("sales_amount", "sum"),
)
rfm["r_score"] = pd.qcut(rfm["recency_days"].rank(method="first"), 4, labels=[4, 3, 2, 1]).astype(int)
rfm["f_score"] = pd.qcut(rfm["frequency"].rank(method="first"), 4, labels=[1, 2, 3, 4]).astype(int)
rfm["m_score"] = pd.qcut(rfm["monetary"].rank(method="first"), 4, labels=[1, 2, 3, 4]).astype(int)


def _segment(row):
    if row.r_score >= 3 and row.f_score >= 3 and row.m_score >= 3:
        return "Champions"
    if row.f_score >= 3 and row.m_score >= 3:
        return "Loyal Customers"
    if row.r_score >= 3 and row.f_score <= 2:
        return "New / Promising"
    if row.r_score <= 2 and row.f_score >= 3:
        return "At Risk"
    return "Needs Attention"


rfm["rfm_segment"] = rfm.apply(_segment, axis=1)
rfm["monetary"] = rfm["monetary"].round(2)
tables["dim_customers"] = tables["dim_customers"].merge(
    rfm[["recency_days", "frequency", "monetary", "rfm_segment"]],
    on="customer_id", how="left",
)
tables["dim_customers"][["recency_days", "frequency", "monetary"]] = \
    tables["dim_customers"][["recency_days", "frequency", "monetary"]].fillna(0)
tables["dim_customers"]["rfm_segment"] = tables["dim_customers"]["rfm_segment"].fillna("No Purchases Yet")

for name, df in tables.items():
    df.to_csv(f"{CSV_DIR}/{name}.csv", index=False)

with pd.ExcelWriter(XLSX_OUT, engine="openpyxl") as writer:
    cover = pd.DataFrame({
        "Table": list(tables.keys()),
        "Rows": [len(df) for df in tables.values()],
        "Role in star schema": [
            "Dimension — customer attributes + precomputed RFM snapshot (join key: customer_id)",
            "Dimension — product attributes (join key: product_id)",
            "Dimension — calendar table, mark as Date table in Power BI (join key: date_id)",
            "Dimension — sales channel lookup (join key: channel_name = fact_sales[channel])",
            "Fact — one row per order line item, grain of the model",
        ],
    })
    cover.to_excel(writer, sheet_name="Model Overview", index=False)
    for name, df in tables.items():
        df.to_excel(writer, sheet_name=name, index=False)

wb = load_workbook(XLSX_OUT)

header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", start_color=NAVY, end_color=NAVY)
body_font = Font(name="Arial", size=10)
title_font = Font(name="Arial", bold=True, size=14, color=NAVY)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    max_col = ws.max_column
    max_row = ws.max_row
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        header_len = len(str(cell.value)) if cell.value else 10
        ws.column_dimensions[get_column_letter(col)].width = max(12, min(28, header_len + 4))
    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).font = body_font
    ws.freeze_panes = "A2"
    if max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

cover_ws = wb["Model Overview"]
cover_ws.insert_rows(1)
cover_ws["A1"] = "NovaCart — Power BI Data Model"
cover_ws["A1"].font = title_font
for col in range(1, cover_ws.max_column + 1):
    cover_ws.cell(row=2, column=col).font = header_font
    cover_ws.cell(row=2, column=col).fill = header_fill

# currency / date number formats on fact_sales and dims
money_cols_by_sheet = {
    "fact_sales": ["unit_price", "sales_amount", "cost_amount", "profit_amount", "shipping_cost"],
    "dim_products": ["unit_price", "cost_price"],
}
for sheet, cols in money_cols_by_sheet.items():
    ws = wb[sheet]
    headers = [c.value for c in ws[1]]
    for col_name in cols:
        if col_name in headers:
            idx = headers.index(col_name) + 1
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=idx).number_format = '"₹"#,##0.00'

wb.save(XLSX_OUT)
print(f"Wrote {XLSX_OUT} with sheets: {wb.sheetnames}")
print(f"Wrote CSVs to {CSV_DIR}/: {list(tables.keys())}")
